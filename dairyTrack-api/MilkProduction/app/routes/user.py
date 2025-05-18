from flask import Blueprint, request, jsonify
from werkzeug.security import generate_password_hash
from app.models.users import User
from app.models.roles import Role  # Tambahkan import Role
from app.database.database import db
from fpdf import FPDF
from flask import send_file
from io import BytesIO
import pandas as pd
from werkzeug.security import check_password_hash


user_bp = Blueprint('user', __name__)

@user_bp.route('/list', methods=['GET'])
def get_all_users():
    try:
        # Ambil semua data pengguna dari database
        users = User.query.all()

        # Format data pengguna menjadi list of dictionaries
        users_list = [{
            "id": user.id,
            "name": user.name,  # Tambahkan name
            "username": user.username,
            "email": user.email,
            "contact": user.contact,
            "religion": user.religion,
            "role_id": user.role_id,
            "birth": user.birth,  # Tambahkan birth
            "token": user.token
        } for user in users]

        return jsonify({"users": users_list}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route('/<int:user_id>', methods=['GET'])
def get_user_by_id(user_id):
    try:
        # Cari user berdasarkan user_id
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Format data pengguna menjadi dictionary
        user_data = {
            "id": user.id,
            "name": user.name,  # Tambahkan name
            "username": user.username,
            "email": user.email,
            "contact": user.contact,
            "religion": user.religion,
            "role_id": user.role_id,
            "birth": user.birth,  # Tambahkan birth
            "token": user.token,
        }

        return jsonify({"user": user_data}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@user_bp.route('/add', methods=['POST'])
def add_user():
    try:
        # Ambil data dari request body
        data = request.get_json()
        name = data.get('name')  # Tambahkan name
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        contact = data.get('contact')
        religion = data.get('religion')
        birth = data.get('birth')  # Tambahkan birth
        role_id = data.get('role_id')  # Pastikan role_id sesuai dengan user
        token = data.get('token', None)

        # Validasi data
        if not name or not username or not email or not password or not role_id:
            return jsonify({"error": "Missing required fields"}), 400

        # Cek apakah username atau email sudah ada
        if User.query.filter_by(username=username).first():
            return jsonify({"error": "Username already exists"}), 400
        if User.query.filter_by(email=email).first():
            return jsonify({"error": "Email already exists"}), 400

        # Validasi role_id
        role = Role.query.get(role_id)
        if not role:
            return jsonify({"error": "Invalid role_id"}), 400

        # Hash password
        hashed_password = generate_password_hash(password)

        # Buat instance User baru
        new_user = User(
            name=name,  # Tambahkan name
            username=username,
            email=email,
            password=hashed_password,
            contact=contact,
            religion=religion,
            birth=birth,  # Tambahkan birth
            role_id=role_id,
            token=token
        )

        # Simpan ke database
        db.session.add(new_user)
        db.session.commit()

        return jsonify({"message": "User added successfully", "user": {
            "name": new_user.name,  # Tambahkan name
            "username": new_user.username,
            "email": new_user.email,
            "contact": new_user.contact,
            "religion": new_user.religion,
            "birth": new_user.birth,  # Tambahkan birth
            "role_id": new_user.role_id
        }}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@user_bp.route('/export/pdf', methods=['GET'])
def export_users_pdf():
    try:
        # Ambil semua data pengguna dari database
        users = User.query.join(Role, User.role_id == Role.id).add_columns(
            User.id, User.name, User.username, User.email, User.contact,
            User.religion, User.birth, Role.name.label('role_name')
        ).all()

        # Buat PDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Tambahkan deskripsi di bagian atas
        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(200, 10, txt="Laporan Data Pengguna", ln=True, align='C')
        pdf.ln(5)
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, txt="Berikut adalah daftar pengguna yang terdaftar dalam sistem.", ln=True, align='C')
        pdf.ln(10)

        # Tambahkan header tabel dengan warna latar belakang
        pdf.set_fill_color(173, 216, 230)  # Warna biru muda (RGB)
        pdf.set_text_color(0, 0, 0)  # Warna teks hitam
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(20, 10, "NO", border=1, align='C', fill=True)  # Ganti ID menjadi NO
        pdf.cell(40, 10, "Name", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Username", border=1, align='C', fill=True)
        pdf.cell(50, 10, "Email", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Role", border=1, align='C', fill=True)
        pdf.ln()

        # Isi data
        pdf.set_font("Arial", size=10)
        for idx, user in enumerate(users, start=1):  # Gunakan enumerate untuk nomor urut
            pdf.cell(20, 10, str(idx), border=1, align='C')  # Tampilkan nomor urut
            pdf.cell(40, 10, user.name, border=1)
            pdf.cell(40, 10, user.username, border=1)
            pdf.cell(50, 10, user.email, border=1)
            pdf.cell(40, 10, user.role_name, border=1)
            pdf.ln()

        # Simpan PDF ke buffer
        buffer = BytesIO()
        pdf.output(buffer)
        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name="users.pdf", mimetype='application/pdf')

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@user_bp.route('/export/excel', methods=['GET'])
def export_users_excel():
    try:
        # Ambil semua data pengguna dari database dengan join ke tabel Role
        users = User.query.join(Role, User.role_id == Role.id).add_columns(
            User.id, User.name, User.username, User.email, User.contact,
            User.religion, User.birth, Role.name.label('role_name')
        ).all()

        # Format data pengguna menjadi list of dictionaries
        users_list = [{
            "NO": idx,  # Tambahkan nomor urut
            "Name": user.name,
            "Username": user.username,
            "Email": user.email,
            "Contact": user.contact,
            "Religion": user.religion,
            "Role": user.role_name,  # Ganti Role ID dengan nama role
            "Birth": user.birth
        } for idx, user in enumerate(users, start=1)]  # Gunakan enumerate untuk nomor urut

        # Buat DataFrame
        df = pd.DataFrame(users_list)

        # Simpan ke buffer dengan pewarnaan
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Users')

            # Akses workbook dan worksheet untuk styling
            workbook = writer.book
            worksheet = writer.sheets['Users']

            # Tambahkan pewarnaan header
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[1]:  # Baris pertama adalah header
                cell.fill = header_fill
                cell.font = header_font

            # Atur lebar kolom secara otomatis
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name="users.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@user_bp.route('/delete/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    try:
        # Cari user berdasarkan user_id
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Hapus user dari database
        db.session.delete(user)
        db.session.commit()

        return jsonify({"message": "User deleted successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    
@user_bp.route('/edit/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    try:
        # Cari user berdasarkan user_id
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Update user data
        data = request.json
        user.name = data.get("name", user.name)
        user.username = data.get("username", user.username)
        user.email = data.get("email", user.email)
        user.contact = data.get("contact", user.contact)
        user.religion = data.get("religion", user.religion)
        user.birth = data.get("birth", user.birth)
        user.role_id = data.get("role_id", user.role_id)

        db.session.commit()

        return jsonify({"message": "User updated successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    


@user_bp.route('/farmers', methods=['GET'])
def get_all_farmers():
    try:
        # Get all users with farmer role by joining User and Role tables
        farmers = User.query.join(Role, User.role_id == Role.id).filter(
            Role.name == 'farmer'  # Assuming 'farmer' is the role name for farmers
        ).add_columns(
            User.id,
            User.name,
            User.username,
            User.email,
            User.contact,
            User.religion,
            User.birth,
            Role.name.label('role_name')
        ).all()

        # Format farmers data into list of dictionaries
        farmers_list = [{
            "id": farmer.id,
            "name": farmer.name,
            "username": farmer.username,
            "email": farmer.email,
            "contact": farmer.contact,
            "religion": farmer.religion,
            "birth": farmer.birth,
            "role": farmer.role_name
        } for farmer in farmers]

        return jsonify({
            "status": "success",
            "message": "Farmers retrieved successfully",
            "total_farmers": len(farmers_list),
            "farmers": farmers_list
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    

@user_bp.route('/reset-password/<int:user_id>', methods=['POST'])
def reset_password(user_id):
    try:
        # Find user by ID
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404
        
        # Get the user's role
        role = Role.query.get(user.role_id)
        if not role:
            return jsonify({"error": "User role not found"}), 404
        
        # Set default password based on role
        default_password = None
        if role.name.lower() == 'farmer':
            default_password = 'farmer123'
        elif role.name.lower() == 'supervisor':
            default_password = 'supervisor123'
        else:
            return jsonify({"error": f"Password reset not supported for role: {role.name}"}), 400
        
        # Hash the default password
        hashed_password = generate_password_hash(default_password)
        
        # Update user password
        user.password = hashed_password
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": f"Password reset successfully for {user.name}",
            "user_id": user.id,
            "role": role.name
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    
@user_bp.route('/change-password/<int:user_id>', methods=['POST'])
def change_password(user_id):
    try:
        data = request.get_json()
        old_password = data.get('old_password')
        new_password = data.get('new_password')

        if not old_password or not new_password:
            return jsonify({"status": "error", "message": "Old and new password are required"}), 400

        user = User.query.get(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404

        # Verifikasi password lama
        if not check_password_hash(user.password, old_password):
            return jsonify({"status": "error", "message": "Old password is incorrect"}), 400

        # Update password baru
        user.password = generate_password_hash(new_password)
        db.session.commit()

        return jsonify({"status": "success", "message": "Password changed successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500