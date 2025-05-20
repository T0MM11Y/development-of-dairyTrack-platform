from flask import Blueprint, request, jsonify
from app.models.milking_sessions import MilkingSession
from app.models.milk_batches import MilkBatch, MilkStatus
from app.models.daily_milk_summary import DailyMilkSummary
from app.database.database import db
from datetime import datetime, date, timedelta
from sqlalchemy import func
from fpdf import FPDF
from flask import send_file
from io import BytesIO
from app.services.notification import check_milk_expiry_and_notify, check_milk_production_and_notify
import pandas as pd

milk_production_bp = Blueprint('milk_production', __name__)

# MilkingSession routes
@milk_production_bp.route('/milking-sessions', methods=['POST'])
def add_milking_session():
    data = request.json
    
    try:
        # Create a new milk batch automatically
        new_batch = MilkBatch(
            batch_number=f"BATCH-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
            total_volume=data['volume'],
            status=MilkStatus.FRESH,
            production_date=datetime.fromisoformat(data.get('milking_time', datetime.utcnow().isoformat())),
            expiry_date=datetime.fromisoformat(data.get('milking_time', datetime.utcnow().isoformat())) + timedelta(hours=8),
            notes=f"Auto-generated batch from milking session. {data.get('notes', '')}"
        )
        
        db.session.add(new_batch)
        db.session.flush()  # This assigns an ID to new_batch without committing
        
        # Now create the milking session with the new batch ID
        new_session = MilkingSession(
            cow_id=data['cow_id'],
            milker_id=data['milker_id'],
            milk_batch_id=new_batch.id,  # Link to the new batch
            volume=data['volume'],
            milking_time=datetime.fromisoformat(data.get('milking_time', datetime.utcnow().isoformat())),
            notes=data.get('notes')
        )
        
        db.session.add(new_session)
        
        # Update the daily milk summary
        session_date = new_session.milking_time.date()
        summary = DailyMilkSummary.query.filter_by(
            cow_id=new_session.cow_id,
            date=session_date
        ).first()
        
        if not summary:
            # Initialize with zero values when creating a new summary
            summary = DailyMilkSummary(
                cow_id=new_session.cow_id,
                date=session_date,
                morning_volume=0,    # Initialize with zero instead of None
                afternoon_volume=0,  # Initialize with zero instead of None
                evening_volume=0,    # Initialize with zero instead of None
                total_volume=0       # Initialize with zero instead of None
            )
            db.session.add(summary)
        
        # Ensure values are initialized if they are None
        if summary.morning_volume is None:
            summary.morning_volume = 0
        if summary.afternoon_volume is None:
            summary.afternoon_volume = 0
        if summary.evening_volume is None:
            summary.evening_volume = 0
        
        # Determine time of day and update corresponding volume
        hour = new_session.milking_time.hour
        if hour < 12:
            summary.morning_volume += float(new_session.volume)
        elif hour < 18:
            summary.afternoon_volume += float(new_session.volume)
        else:
            summary.evening_volume += float(new_session.volume)
            
        # Ensure total_volume is not None
        if summary.total_volume is None:
            summary.total_volume = 0
        summary.total_volume = summary.morning_volume + summary.afternoon_volume + summary.evening_volume
        
        db.session.commit()
        #cek evening kosong apatidak
        if summary.evening_volume != 0 or summary.afternoon_volume != 0:
            check_milk_production_and_notify()
            check_milk_expiry_and_notify()

        return jsonify({
            "success": True, 
            "message": "Milking session added successfully with new batch", 
            "id": new_session.id,
            "batch_id": new_batch.id
        }), 201
    
    except Exception as e:
        db.session.rollback()   
        return jsonify({"success": False, "error": str(e)}), 400

@milk_production_bp.route('/milking-sessions', methods=['GET'])
def get_milking_sessions():
    sessions = MilkingSession.query.all()
    result = []
    
    for session in sessions:
        result.append({
            "id": session.id,
            "cow_id": session.cow_id,
            "cow_name": session.cow.name if session.cow else None,
            "milker_id": session.milker_id,
            "milker_name": session.milker.name if session.milker else None,
            "milk_batch_id": session.milk_batch_id,
            "volume": session.volume,
            "milking_time": session.milking_time.isoformat(),
            "notes": session.notes
        })
    
    return jsonify(result), 200


@milk_production_bp.route('/milk-batches', methods=['GET'])
def get_milk_batches():
    batches = MilkBatch.query.all()
    result = []
    
    for batch in batches:
        result.append({
            "id": batch.id,
            "batch_number": batch.batch_number,
            "total_volume": batch.total_volume,
            "status": batch.status.value,
            "production_date": batch.production_date.isoformat(),
            "expiry_date": batch.expiry_date.isoformat() if batch.expiry_date else None,
            "notes": batch.notes
        })
    
    return jsonify(result), 200

@milk_production_bp.route('/daily-summaries', methods=['GET'])
def get_daily_summaries():
    try:
        # Get query parameters
        cow_id = request.args.get('cow_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Initialize query
        query = DailyMilkSummary.query
        
        # Apply filters
        if cow_id:
            try:
                cow_id = int(cow_id)
                query = query.filter_by(cow_id=cow_id)
            except ValueError:
                return jsonify({
                    "success": False,
                    "error": "Invalid cow_id format. Must be an integer."
                }), 400
        
        # Validate and apply date filters
        try:
            if start_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(DailyMilkSummary.date >= start_date)
            
            if end_date:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(DailyMilkSummary.date <= end_date)
                
            if start_date and end_date and start_date > end_date:
                return jsonify({
                    "success": False,
                    "error": "start_date cannot be later than end_date"
                }), 400
                
        except ValueError:
            return jsonify({
                "success": False,
                "error": "Invalid date format. Use YYYY-MM-DD"
            }), 400
        
        # Execute query and format results
        summaries = query.order_by(DailyMilkSummary.date.desc()).all()
        result = []
        
        for summary in summaries:
            result.append({
                "id": summary.id,
                "cow_id": summary.cow_id,
                "cow_name": summary.cow.name if summary.cow else None,
                "date": summary.date.isoformat(),
                "morning_volume": float(summary.morning_volume or 0),
                "afternoon_volume": float(summary.afternoon_volume or 0),
                "evening_volume": float(summary.evening_volume or 0),
                "total_volume": float(summary.total_volume or 0)
            })
        
        return jsonify({
            "success": True,
            "summaries": result,
            "total_records": len(result)
        }), 200
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"An error occurred while fetching daily summaries: {str(e)}"
        }), 500


@milk_production_bp.route('/export/pdf', methods=['GET'])
def export_milking_sessions_pdf():
    try:
        sessions = MilkingSession.query.all()

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(200, 10, txt="Laporan Data Milking Sessions", ln=True, align='C')
        pdf.ln(5)
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, txt="Daftar sesi pemerahan sapi.", ln=True, align='C')
        pdf.ln(10)

        pdf.set_fill_color(173, 216, 230)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(10, 10, "NO", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Cow", border=1, align='C', fill=True)
        pdf.cell(40, 10, "Milker", border=1, align='C', fill=True)
        pdf.cell(25, 10, "Session", border=1, align='C', fill=True)
        pdf.cell(25, 10, "Volume", border=1, align='C', fill=True)
        pdf.cell(45, 10, "Milking Time", border=1, align='C', fill=True)
        pdf.ln()

        pdf.set_font("Arial", size=10)
        for idx, session in enumerate(sessions, start=1):
            cow_info = f"{session.cow_id} - {session.cow.name}" if session.cow else str(session.cow_id)
            milker_info = f"{session.milker_id} - {session.milker.name}" if session.milker else str(session.milker_id)
            # Tentukan sesi berdasarkan jam
            hour = session.milking_time.hour
            if hour < 12:
                sesi = "Pagi"
            elif hour < 18:
                sesi = "Siang"
            else:
                sesi = "Sore"
            pdf.cell(10, 10, str(idx), border=1, align='C')
            pdf.cell(40, 10, cow_info, border=1)
            pdf.cell(40, 10, milker_info, border=1)
            pdf.cell(25, 10, sesi, border=1, align='C')
            pdf.cell(25, 10, str(session.volume), border=1)
            pdf.cell(45, 10, session.milking_time.strftime('%Y-%m-%d %H:%M'), border=1)
            pdf.ln()

        buffer = BytesIO()
        pdf.output(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="milking_sessions.pdf", mimetype='application/pdf')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@milk_production_bp.route('/export/excel', methods=['GET'])
def export_milking_sessions_excel():
    try:
        sessions = MilkingSession.query.all()
        sessions_list = []
        for idx, session in enumerate(sessions, start=1):
            hour = session.milking_time.hour
            if hour < 12:
                sesi = "Pagi"
            elif hour < 18:
                sesi = "Siang"
            else:
                sesi = "Sore"
            sessions_list.append({
                "NO": idx,
                "Cow": f"{session.cow_id} - {session.cow.name}" if session.cow else str(session.cow_id),
                "Milker": f"{session.milker_id} - {session.milker.name}" if session.milker else str(session.milker_id),
                "Session": sesi,
                "Volume": session.volume,
                "Milking Time": session.milking_time.strftime('%Y-%m-%d %H:%M')
            })

        df = pd.DataFrame(sessions_list)
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='MilkingSessions')
            workbook = writer.book
            worksheet = writer.sheets['MilkingSessions']
            from openpyxl.styles import Font, PatternFill
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="milking_sessions.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ...existing code...

@milk_production_bp.route('/milking-sessions/<int:session_id>', methods=['DELETE'])
def delete_milking_session(session_id):
    try:
        session = MilkingSession.query.get(session_id)
        if not session:
            return jsonify({"success": False, "error": "Milking session not found"}), 404
        
        # Store information before deletion for summary update
        cow_id = session.cow_id
        session_date = session.milking_time.date()
        volume = session.volume
        hour = session.milking_time.hour
        milk_batch_id = session.milk_batch_id
        
        # Delete the milking session
        db.session.delete(session)
        db.session.flush()
        
        # Check if we need to update the milk batch
        batch = MilkBatch.query.get(milk_batch_id)
        if batch:
            # Reduce the batch volume
            batch.total_volume -= volume
            
            # If batch has no more volume or no more sessions, delete it
            remaining_sessions = MilkingSession.query.filter_by(milk_batch_id=milk_batch_id).count()
            if batch.total_volume <= 0 or remaining_sessions == 0:
                db.session.delete(batch)
        
        # Update the daily summary
        summary = DailyMilkSummary.query.filter_by(cow_id=cow_id, date=session_date).first()
        if summary:
            # Subtract the volume from the appropriate time period
            if hour < 12:
                summary.morning_volume -= volume
                if summary.morning_volume < 0:  # Safeguard against negative values
                    summary.morning_volume = 0
            elif hour < 18:
                summary.afternoon_volume -= volume
                if summary.afternoon_volume < 0:
                    summary.afternoon_volume = 0
            else:
                summary.evening_volume -= volume
                if summary.evening_volume < 0:
                    summary.evening_volume = 0
                    
            # Update total volume
            summary.total_volume = summary.morning_volume + summary.afternoon_volume + summary.evening_volume
            
            # If there's no more milk recorded for this cow on this day, delete the summary
            if summary.total_volume <= 0:
                db.session.delete(summary)
        
        db.session.commit()
        return jsonify({"success": True, "message": "Milking session deleted successfully"}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400

# ...existing code...

@milk_production_bp.route('/milking-sessions/<int:session_id>', methods=['PUT'])
def update_milking_session(session_id):
    try:
        session = MilkingSession.query.get(session_id)
        if not session:
            return jsonify({"success": False, "error": "Milking session not found"}), 404
        
        data = request.json
        
        # Store old values for calculations
        old_volume = session.volume
        old_milking_time = session.milking_time
        old_hour = old_milking_time.hour
        old_date = old_milking_time.date()
        old_cow_id = session.cow_id
        
        # Store the new values for calculations
        new_volume = float(data.get('volume', old_volume))
        new_milking_time = datetime.fromisoformat(data.get('milking_time', old_milking_time.isoformat()))
        new_hour = new_milking_time.hour
        new_date = new_milking_time.date()
        new_cow_id = int(data.get('cow_id', old_cow_id))
        
        # Update the session fields
        session.cow_id = new_cow_id
        session.milker_id = data.get('milker_id', session.milker_id)
        session.volume = new_volume
        session.milking_time = new_milking_time
        session.notes = data.get('notes', session.notes)
        
        # Update the milk batch if volume changed
        if old_volume != new_volume:
            batch = MilkBatch.query.get(session.milk_batch_id)
            if batch:
                batch.total_volume = batch.total_volume - old_volume + new_volume
                
                # Update expiry date if milking time changed
                if old_milking_time != new_milking_time:
                    batch.production_date = new_milking_time
                    batch.expiry_date = new_milking_time + timedelta(hours=8)
                
        # Handle daily milk summary updates
        
        # 1. If date or cow changed, we need to update two summaries
        if old_date != new_date or old_cow_id != new_cow_id:
            # First, update the old summary (subtract volume)
            old_summary = DailyMilkSummary.query.filter_by(
                cow_id=old_cow_id,
                date=old_date
            ).first()
            
            if old_summary:
                # Subtract the old volume from the appropriate time period
                if old_hour < 12:
                    old_summary.morning_volume -= old_volume
                    if old_summary.morning_volume < 0:
                        old_summary.morning_volume = 0
                elif old_hour < 18:
                    old_summary.afternoon_volume -= old_volume
                    if old_summary.afternoon_volume < 0:
                        old_summary.afternoon_volume = 0
                else:
                    old_summary.evening_volume -= old_volume
                    if old_summary.evening_volume < 0:
                        old_summary.evening_volume = 0
                        
                # Update total volume
                old_summary.total_volume = old_summary.morning_volume + old_summary.afternoon_volume + old_summary.evening_volume
                
                # If there's no more milk recorded for this cow on this day, delete the summary
                if old_summary.total_volume <= 0:
                    db.session.delete(old_summary)
            
            # Then, find or create new summary and add volume
            new_summary = DailyMilkSummary.query.filter_by(
                cow_id=new_cow_id,
                date=new_date
            ).first()
            
            if not new_summary:
                new_summary = DailyMilkSummary(
                    cow_id=new_cow_id,
                    date=new_date,
                    morning_volume=0,
                    afternoon_volume=0,
                    evening_volume=0,
                    total_volume=0
                )
                db.session.add(new_summary)
            
            # Add the new volume to the appropriate time period
            if new_hour < 12:
                new_summary.morning_volume += new_volume
            elif new_hour < 18:
                new_summary.afternoon_volume += new_volume
            else:
                new_summary.evening_volume += new_volume
                
            # Update total volume
            new_summary.total_volume = new_summary.morning_volume + new_summary.afternoon_volume + new_summary.evening_volume
        
        # 2. If only volume or time of day changed but date and cow are the same
        elif old_volume != new_volume or (old_hour < 12 and new_hour >= 12) or (old_hour < 18 and new_hour >= 18) or (old_hour >= 18 and new_hour < 18):
            summary = DailyMilkSummary.query.filter_by(
                cow_id=new_cow_id,
                date=new_date
            ).first()
            
            if summary:
                # Remove old volume from old time period
                if old_hour < 12:
                    summary.morning_volume -= old_volume
                    if summary.morning_volume < 0:
                        summary.morning_volume = 0
                elif old_hour < 18:
                    summary.afternoon_volume -= old_volume
                    if summary.afternoon_volume < 0:
                        summary.afternoon_volume = 0
                else:
                    summary.evening_volume -= old_volume
                    if summary.evening_volume < 0:
                        summary.evening_volume = 0
                
                # Add new volume to new time period
                if new_hour < 12:
                    summary.morning_volume += new_volume
                elif new_hour < 18:
                    summary.afternoon_volume += new_volume
                else:
                    summary.evening_volume += new_volume
                
                # Update total
                summary.total_volume = summary.morning_volume + summary.afternoon_volume + summary.evening_volume
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Milking session updated successfully",
            "id": session.id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    


@milk_production_bp.route('/export/daily-summaries/pdf', methods=['GET'])
def export_daily_summaries_pdf():
    try:
        # Get query parameters
        cow_id = request.args.get('cow_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Initialize query
        query = DailyMilkSummary.query
        
        # Apply filters
        if cow_id:
            try:
                cow_id = int(cow_id)
                query = query.filter_by(cow_id=cow_id)
            except ValueError:
                return jsonify({
                    "success": False,
                    "error": "Invalid cow_id format. Must be an integer."
                }), 400
        
        # Validate and apply date filters
        try:
            if start_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(DailyMilkSummary.date >= start_date)
            
            if end_date:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(DailyMilkSummary.date <= end_date)
                
            if start_date and end_date and start_date > end_date:
                return jsonify({
                    "success": False,
                    "error": "start_date cannot be later than end_date"
                }), 400
                
        except ValueError:
            return jsonify({
                "success": False,
                "error": "Invalid date format. Use YYYY-MM-DD"
            }), 400
        
        # Execute query and get summaries
        summaries = query.order_by(DailyMilkSummary.date.desc()).all()
        
        # Create PDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(200, 10, txt="Laporan Produksi Susu Harian", ln=True, align='C')
        pdf.ln(5)
        
        # Add filter information
        pdf.set_font("Arial", size=10)
        filter_text = "Filter: "
        if cow_id:
            cow = summaries[0].cow.name if summaries and summaries[0].cow else f"Cow ID: {cow_id}"
            filter_text += f"Sapi: {cow}, "
        if start_date:
            filter_text += f"Dari: {start_date.strftime('%Y-%m-%d')}, "
        if end_date:
            filter_text += f"Sampai: {end_date.strftime('%Y-%m-%d')}, "
        
        if filter_text == "Filter: ":
            filter_text += "Semua data"
        else:
            filter_text = filter_text[:-2]  # Remove last comma and space
            
        pdf.cell(200, 10, txt=filter_text, ln=True, align='C')
        pdf.ln(10)

        # Create table headers
        pdf.set_fill_color(173, 216, 230)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(10, 10, "NO", border=1, align='C', fill=True)
        pdf.cell(50, 10, "Sapi", border=1, align='C', fill=True)
        pdf.cell(30, 10, "Tanggal", border=1, align='C', fill=True)
        pdf.cell(25, 10, "Pagi", border=1, align='C', fill=True)
        pdf.cell(25, 10, "Siang", border=1, align='C', fill=True)
        pdf.cell(25, 10, "Sore", border=1, align='C', fill=True)
        pdf.cell(25, 10, "Total", border=1, align='C', fill=True)
        pdf.ln()

        # Add data rows
        pdf.set_font("Arial", size=10)
        for idx, summary in enumerate(summaries, start=1):
            cow_info = f"{summary.cow_id} - {summary.cow.name}" if summary.cow else str(summary.cow_id)
            
            pdf.cell(10, 10, str(idx), border=1, align='C')
            pdf.cell(50, 10, cow_info, border=1)
            pdf.cell(30, 10, summary.date.strftime('%Y-%m-%d'), border=1, align='C')
            pdf.cell(25, 10, str(round(float(summary.morning_volume or 0), 2)), border=1, align='R')
            pdf.cell(25, 10, str(round(float(summary.afternoon_volume or 0), 2)), border=1, align='R')
            pdf.cell(25, 10, str(round(float(summary.evening_volume or 0), 2)), border=1, align='R')
            pdf.cell(25, 10, str(round(float(summary.total_volume or 0), 2)), border=1, align='R')
            pdf.ln()

        # Add totals row
        total_morning = sum(float(s.morning_volume or 0) for s in summaries)
        total_afternoon = sum(float(s.afternoon_volume or 0) for s in summaries)
        total_evening = sum(float(s.evening_volume or 0) for s in summaries)
        total_all = sum(float(s.total_volume or 0) for s in summaries)
        
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(90, 10, "TOTAL", border=1, align='C', fill=True)
        pdf.cell(25, 10, str(round(total_morning, 2)), border=1, align='R', fill=True)
        pdf.cell(25, 10, str(round(total_afternoon, 2)), border=1, align='R', fill=True)
        pdf.cell(25, 10, str(round(total_evening, 2)), border=1, align='R', fill=True)
        pdf.cell(25, 10, str(round(total_all, 2)), border=1, align='R', fill=True)

        # Output PDF file
        buffer = BytesIO()
        pdf.output(buffer)
        buffer.seek(0)
        
        filename = "daily_milk_production.pdf"
        if start_date and end_date:
            filename = f"milk_production_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.pdf"
        elif cow_id:
            filename = f"milk_production_cow_{cow_id}.pdf"
            
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')
    
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@milk_production_bp.route('/export/daily-summaries/excel', methods=['GET'])
def export_daily_summaries_excel():
    try:
        # Get query parameters
        cow_id = request.args.get('cow_id')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Initialize query
        query = DailyMilkSummary.query
        
        # Apply filters
        if cow_id:
            try:
                cow_id = int(cow_id)
                query = query.filter_by(cow_id=cow_id)
            except ValueError:
                return jsonify({
                    "success": False,
                    "error": "Invalid cow_id format. Must be an integer."
                }), 400
        
        # Validate and apply date filters
        try:
            if start_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(DailyMilkSummary.date >= start_date)
            
            if end_date:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(DailyMilkSummary.date <= end_date)
                
            if start_date and end_date and start_date > end_date:
                return jsonify({
                    "success": False,
                    "error": "start_date cannot be later than end_date"
                }), 400
                
        except ValueError:
            return jsonify({
                "success": False,
                "error": "Invalid date format. Use YYYY-MM-DD"
            }), 400
        
        # Execute query and get summaries
        summaries = query.order_by(DailyMilkSummary.date.desc()).all()
        
        # Prepare Excel data
        summaries_list = []
        for idx, summary in enumerate(summaries, start=1):
            cow_info = f"{summary.cow_id} - {summary.cow.name}" if summary.cow else str(summary.cow_id)
            
            summaries_list.append({
                "NO": idx,
                "Sapi": cow_info,
                "Tanggal": summary.date.strftime('%Y-%m-%d'),
                "Produksi Pagi": float(summary.morning_volume or 0),
                "Produksi Siang": float(summary.afternoon_volume or 0),
                "Produksi Sore": float(summary.evening_volume or 0),
                "Total Produksi": float(summary.total_volume or 0)
            })
        
        # Add total row
        if summaries_list:
            total_morning = sum(float(s.morning_volume or 0) for s in summaries)
            total_afternoon = sum(float(s.afternoon_volume or 0) for s in summaries)
            total_evening = sum(float(s.evening_volume or 0) for s in summaries)
            total_all = sum(float(s.total_volume or 0) for s in summaries)
            
            summaries_list.append({
                "NO": "",
                "Sapi": "TOTAL",
                "Tanggal": "",
                "Produksi Pagi": total_morning,
                "Produksi Siang": total_afternoon,
                "Produksi Sore": total_evening,
                "Total Produksi": total_all
            })

        # Create DataFrame and Excel file
        df = pd.DataFrame(summaries_list)
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='DailyMilkProduction')
            workbook = writer.book
            worksheet = writer.sheets['DailyMilkProduction']
            
            # Apply styling
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)
            total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            total_font = Font(bold=True)
            
            # Style headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style totals row if it exists
            if len(summaries_list) > 1:
                for cell in worksheet[len(summaries_list) + 1]:  # +1 for header row
                    cell.fill = total_fill
                    cell.font = total_font
            
            # Auto-adjust column widths
            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value) or "") for cell in column_cells)
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width
        
        # Prepare file for download
        buffer.seek(0)
        
        filename = "daily_milk_production.xlsx"
        if start_date and end_date:
            filename = f"milk_production_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
        elif cow_id:
            filename = f"milk_production_cow_{cow_id}.xlsx"
            
        return send_file(
            buffer, 
            as_attachment=True, 
            download_name=filename, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    

@milk_production_bp.route('/check-production', methods=['POST'])
def check_production():
    try:
        production_result = check_milk_production_and_notify()
        return jsonify({
            'success': True,
            'message': 'Production check completed',
            'notifications': production_result
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@milk_production_bp.route('/check-expiry', methods=['POST'])
def check_expiry():
    try:
        expiry_result = check_milk_expiry_and_notify()
        return jsonify({
            'success': True,
            'message': 'Expiry check completed',
            'notifications': expiry_result
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500